import openpyxl as xl
from openpyxl.chart import BarChart, Reference

ws = xl.load_workbook("transactions.xlsx")
sheet = ws["Sheet1"]
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    new_price = cell.value * 0.9
    new_price_cell = sheet.cell(row, 4)
    new_price_cell.value = new_price

reference = Reference(sheet,
                      min_row=2,
                      max_row=sheet.max_row,
                      min_col=4,
                      max_col=4)
barchart = BarChart()
barchart.add_data(reference)
sheet.add_chart(barchart, "e2")

ws.save("transactions3.xlsx")
